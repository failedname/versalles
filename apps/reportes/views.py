from django.shortcuts import render
from django.views.generic import TemplateView
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from openpyxl import Workbook
from django.http import JsonResponse, HttpResponse
from apps.ventas.models import Detalle_FacturaReal, detalleUser, detalleRemison, Categoria, Producto

import json


@method_decorator(login_required, name='dispatch')
class ventasReport(TemplateView):
    template_name = 'reportes/ventas.html'


def report_ventas(request):
    if 'vivero' in request.session:
        fechas = json.loads(request.body.decode('utf-8'))

        data = Detalle_FacturaReal.objects.extra(
            select={'total': 'SELECT sum(ventas_detalle_facturareal.val_neto)  FROM ventas_detalle_facturareal WHERE ventas_detalle_facturareal.factura_id = ventas_facturareal.codigo',
                    'iva': 'SELECT sum(ventas_detalle_facturareal.iva) FROM ventas_detalle_facturareal WHERE ventas_detalle_facturareal.factura_id = ventas_facturareal.codigo'
                    }).select_related(
            'factura',
            'factura__estado',
            'factura__vivero',
            'factura__cliente').filter(
            factura__vivero_id=request.session['vivero'], factura__fecha__gte=fechas['start'], factura__fecha__lte=fechas['end']).distinct(
            'factura_id')
        fact = [{
            'id': res.factura.pk,
            'codigo': res.factura.codigo,
            'fecha': res.factura.fecha,
            'identificacion': res.factura.cliente.nit_cc,
            'nombre': res.factura.cliente.nombre,
            'estado': res.factura.estado.estado,
            'totaliva': res.iva,
            'total': res.total
        }for res in data]
        return JsonResponse({'data': fact}, safe=True)
    else:
        vivero = detalleUser.objects.get(usuario__pk=request.user.id)
        request.session['vivero'] = vivero.vivero.id
        fechas = json.loads(request.body.decode('utf-8'))

        data = Detalle_FacturaReal.objects.extra(
            select={'total': 'SELECT sum(ventas_detalle_facturareal.val_neto)  FROM ventas_detalle_facturareal WHERE ventas_detalle_facturareal.factura_id = ventas_facturareal.codigo',
                    'iva': 'SELECT sum(ventas_detalle_facturareal.iva) FROM ventas_detalle_facturareal WHERE ventas_detalle_facturareal.factura_id = ventas_facturareal.codigo'
                    }).select_related(
            'factura',
            'factura__estado',
            'factura__vivero',
            'factura__cliente').filter(
            factura__vivero_id=request.session['vivero'], factura__fecha__gte=fechas['start'], factura__fecha__lte=fechas['end'], factura__estado__estado='cerrada').distinct(
            'factura_id')
        fact = [{
            'id': res.factura.pk,
            'codigo': res.factura.codigo,
            'fecha': res.factura.fecha,
            'identificacion': res.factura.cliente.nit_cc,
            'nombre': res.factura.cliente.nombre,
            'estado': res.factura.estado.estado,
            'totaliva': res.iva,
            'total': res.total
        }for res in data]
        return JsonResponse({'data': fact}, safe=True)


def export_ventas(request, start, end):
    if 'vivero' in request.session:
        data = Detalle_FacturaReal.objects.extra(
            select={'total': 'SELECT sum(ventas_detalle_facturareal.val_neto)  FROM ventas_detalle_facturareal WHERE ventas_detalle_facturareal.factura_id = ventas_facturareal.codigo',
                    'iva': 'SELECT sum(ventas_detalle_facturareal.iva) FROM ventas_detalle_facturareal WHERE ventas_detalle_facturareal.factura_id = ventas_facturareal.codigo'
                    }).select_related(
            'factura',
            'factura__estado',
            'factura__vivero',
            'factura__cliente').filter(
            factura__vivero_id=request.session['vivero'], factura__fecha__gte=start, factura__fecha__lte=end).distinct(
            'factura_id')
        fact = [{
            'id': res.factura.pk,
            'codigo': res.factura.codigo,
            'fecha': str(res.factura.fecha),
            'identificacion': res.factura.cliente.nit_cc,
            'nombre': res.factura.cliente.nombre,
            'estado': res.factura.estado.estado,
            'totaliva': res.iva,
            'total': res.total
        }for res in data]
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE VENTAS'
        ws.merge_cells('B1:E1')
        # Creamos los encabezados desde la celda B3 hasta la E3
        ws['B3'] = 'FACTURA'
        ws['C3'] = 'FECHA'
        ws['D3'] = 'IDENTIFICACIÓN'
        ws['E3'] = 'NOMBRE'
        ws['F3'] = 'IVA'
        ws['G3'] = 'TOTAL'
        cont = 4
        for ventas in data:
            ws.cell(row=cont, column=2).value = ventas.factura.pk
            ws.cell(row=cont, column=3).value = ventas.factura.fecha
            ws.cell(row=cont, column=4).value = ventas.factura.cliente.nit_cc
            ws.cell(row=cont, column=5).value = ventas.factura.cliente.nombre
            ws.cell(row=cont, column=6).value = ventas.iva
            ws.cell(row=cont, column=7).value = ventas.total
            cont = cont + 1
        nombre_archivo = "ReporteVentasExcel.xlsx"
        # Definimos que el tipo de respuesta a devolver es un archivo de microsoft
        # excel
        response = HttpResponse(content_type="application/ms-excel")
        contenido = "attachment; filename={0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)
        return response
        return JsonResponse({'data': fact}, safe=True)
    # else:
    #     vivero = detalleUser.objects.get(usuario__pk=request.user.id)
    #     request.session['vivero'] = vivero.vivero.id
    #     fechas = json.loads(request.body)
    #
    #     data = Detalle_FacturaReal.objects.extra(
    #         select={'total': 'SELECT sum(ventas_detalle_facturareal.val_neto)  FROM ventas_detalle_facturareal WHERE ventas_detalle_facturareal.factura_id = ventas_facturareal.codigo',
    #                 'iva': 'SELECT sum(ventas_detalle_facturareal.iva) FROM ventas_detalle_facturareal WHERE ventas_detalle_facturareal.factura_id = ventas_facturareal.codigo'
    #                 }).select_related(
    #         'factura',
    #         'factura__estado',
    #         'factura__vivero',
    #         'factura__cliente').filter(
    #         factura__vivero_id=request.session['vivero'], factura__fecha__gte=fechas['start'], factura__fecha__lte=fechas['end'], factura__estado__estado='cerrada').distinct(
    #         'factura_id')
    #     fact = [{
    #         'id': res.factura.pk,
    #         'codigo': res.factura.codigo,
    #         'fecha': str(res.factura.fecha),
    #         'identificacion': res.factura.cliente.nit_cc,
    #         'nombre': res.factura.cliente.nombre,
    #         'estado': res.factura.estado.estado,
    #         'totaliva': res.iva,
    #         'total': res.total
    #     }for res in data]
    #     return JsonResponse({'data': fact}, safe=True)


@method_decorator(login_required, name='dispatch')
class report_remisiones(TemplateView):
    template_name = 'reportes/remisiones.html'

    def post(self, request, *args, **kwargs):
        if 'vivero' in request.session:
            fechas = json.loads(request.body.decode('utf-8'))

            data = detalleRemison.objects.extra(
                select={'total': 'SELECT sum(ventas_detalleremison.val_neto)  FROM ventas_detalleremison WHERE ventas_detalleremison.remision_id = ventas_remision.id',
                        'iva': 'SELECT sum(ventas_detalleremison.iva) FROM ventas_detalleremison WHERE ventas_detalleremison.remision_id = ventas_remision.id'
                        }).select_related(
                'remision',
                'remision__estado',
                'remision__vivero',
                'remision__cliente').filter(
                remision__vivero_id=request.session['vivero'], remision__fecha__gte=fechas['start'], remision__fecha__lte=fechas['end'], remision__estado__estado='cerrada').distinct(
                'remision_id')
            fact = [{
                'codigo': res.remision.pk,
                'fecha': str(res.remision.fecha),
                'identificacion': res.remision.cliente.nit_cc,
                'nombre': res.remision.cliente.nombre,
                'estado': res.remision.estado.estado,
                'totaliva': res.iva,
                'total': res.total
            }for res in data]
            return JsonResponse({'data': fact}, safe=True)
        else:
            vivero = detalleUser.objects.get(usuario__pk=request.user.id)
            request.session['vivero'] = vivero.vivero.id
            fechas = json.loads(request.body.decode('utf-8'))

            data = detalleRemison.objects.extra(
                select={'total': 'SELECT sum(ventas_detalleremison.val_neto)  FROM ventas_detalleremison WHERE ventas_detalleremison.remision_id = ventas_remision.id',
                        'iva': 'SELECT sum(ventas_detalleremison.iva) FROM ventas_detalleremison WHERE ventas_detalleremison.remision_id = ventas_remision.id'
                        }).select_related(
                'remision',
                'remision__estado',
                'remision__vivero',
                'remision__cliente').filter(
                remision__vivero_id=request.session['vivero'], remision__fecha__gte=fechas['start'], remision__fecha__lte=fechas['end'], remision__estado__estado='cerrada').distinct(
                'remision_id')
            fact = [{
                'codigo': res.remision.pk,
                'fecha': str(res.remision.fecha),
                'identificacion': res.remision.cliente.nit_cc,
                'nombre': res.remision.cliente.nombre,
                'estado': res.remision.estado.estado,
                'totaliva': res.iva,
                'total': res.total
            }for res in data]
            return JsonResponse({'data': fact}, safe=True)


class ReportCat(TemplateView):
    template_name = 'reportes/categoria.html'

    def get(self, request, *args, **kwargs):
        cat = Categoria.objects.all()
        return render(request, self.template_name, {'data': cat})

    def post(self, request, *args, **kwargs):
        if 'vivero' in request.session:
            fechas = json.loads(request.body.decode('utf-8'))

            data = Producto.objects.raw(''' SELECT
            ventas_producto.nombre,
            sum(ventas_detalle_facturareal.cantidad) AS "cantidad",
            sum(ventas_detalle_facturareal.val_neto) AS "total_venta",
            sum(ventas_detalle_facturareal.iva) AS "iva",
            sum(ventas_producto.valor_real_compra * cantidad) AS "valor_compra",
            ventas_producto.id
            FROM
            ventas_producto
            JOIN ventas_detalle_facturareal
            ON ventas_producto.id = ventas_detalle_facturareal.producto_id
            JOIN "public"."ventas_facturareal"
            ON ventas_detalle_facturareal.factura_id = ventas_facturareal.codigo
            WHERE
            ventas_facturareal.fecha BETWEEN %s AND %s AND ventas_facturareal.vivero_id=%s AND ventas_producto.id_categoria_id=%s  AND (ventas_facturareal.estado_id=1 OR ventas_facturareal.estado_id=3 ) GROUP BY
            ventas_producto.id ''', [fechas['start'], fechas['end'], request.session['vivero'], fechas['cate']])
            fact = [{
                'id': res.pk,
                'nombre': res.nombre,
                'cantidad': res.cantidad,
                'totalventa': res.total_venta,
                'iva': res.iva,
                'totalcompra': res.valor_compra
            }for res in data]
            return JsonResponse({'data': fact}, safe=True)
        else:
            vivero = detalleUser.objects.get(usuario__pk=request.user.id)
            request.session['vivero'] = vivero.vivero.id
            fechas = json.loads(request.body.decode('utf-8'))

            data = Producto.objects.raw(''' SELECT
            ventas_producto.nombre,
            sum(ventas_detalle_facturareal.cantidad) AS "cantidad",
            sum(ventas_detalle_facturareal.val_neto) AS "total_venta",
            sum(ventas_detalle_facturareal.iva) AS "iva",
            sum(ventas_producto.valor_real_compra * cantidad) AS "valor_compra",
            ventas_producto.id
            FROM
            ventas_producto
            JOIN ventas_detalle_facturareal
            ON ventas_producto.id = ventas_detalle_facturareal.producto_id
            JOIN "public"."ventas_facturareal"
            ON ventas_detalle_facturareal.factura_id = ventas_facturareal.codigo
            WHERE
            (ventas_facturareal.fecha BETWEEN %s AND %s) and (ventas_facturareal.vivero_id=%s) AND ventas_producto.id_categoria_id=%s AND (ventas_facturareal.estado_id=1 OR ventas_facturareal.estado_id=3 )
            GROUP BY
            ventas_producto.id ''', [fechas['start'], fechas['end'], request.session['vivero'], fechas['cate']])
            fact = [{
                'id': res.pk,
                'nombre': res.nombre,
                'cantidad': res.cantidad,
                'totalventa': res.total_venta,
                'iva': res.iva,
                'totalcompra': res.valor_compra
            }for res in data]
            return JsonResponse({'data': fact}, safe=True)
